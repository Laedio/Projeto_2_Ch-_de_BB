import openpyxl
import os
from datetime import datetime

# 1. Definimos o nome EXATO do arquivo que você criou na pasta
NOME_EXCEL = 'lista_ayla.xlsx'

def inicializar_banco():
    # Apenas verifica se você criou o arquivo como combinamos
    if not os.path.exists(NOME_EXCEL):
        print(f"ERRO: O arquivo {NOME_EXCEL} não foi encontrado na pasta!")
    else:
        print("Planilha do Excel detectada e pronta para uso!")

def salvar_confirmacao(nome, fralda, mimo, presenca):
    book = openpyxl.load_workbook(NOME_EXCEL)
    folha = book.active
    
    data_atual = datetime.now().strftime('%d/%m/%Y %H:%M')
    nome = nome.strip().title()
    mimo = mimo.strip() if mimo else "Nenhum"

    # 1. Encontrar a primeira linha onde o NOME (Coluna B) está vazio
    linha_vazia = 2 
    while folha.cell(row=linha_vazia, column=2).value is not None:
        linha_vazia += 1

    # 2. Escrever os dados nas colunas de A até E
    folha.cell(row=linha_vazia, column=1).value = data_atual
    folha.cell(row=linha_vazia, column=2).value = nome
    folha.cell(row=linha_vazia, column=3).value = fralda
    folha.cell(row=linha_vazia, column=4).value = mimo
    folha.cell(row=linha_vazia, column=5).value = presenca # Coluna de confirmação
    
    # 3. Salvar o arquivo (Removida a lógica de G1 e H1)
    book.save(NOME_EXCEL)
    print(f"Registro de {nome} salvo com sucesso!")
    if not os.path.exists(NOME_EXCEL):
        return []

    book = openpyxl.load_workbook(NOME_EXCEL)
    folha = book.active
    
    lista_convidados = []
    
    # min_row=2 pula o cabeçalho
    for linha in folha.iter_rows(min_row=2, values_only=True):
        # linha[0]=Data, [1]=Nome, [2]=Fralda, [3]=Mimo, [4]=Presença
        if linha[1]: 
            convidado = {
                'data': linha[0],
                'nome': linha[1],
                'fralda': linha[2],
                'mimo': linha[3],
                'presenca': linha[4] # Lendo a nova coluna!
            }
            lista_convidados.append(convidado)
            
    return lista_convidados

def ler_confirmacoes():
    if not os.path.exists(NOME_EXCEL):
        return []

    book = openpyxl.load_workbook(NOME_EXCEL)
    folha = book.active
    
    lista_convidados = []
    
    # min_row=2 pula o cabeçalho colorido
    for linha in folha.iter_rows(min_row=2, values_only=True):
        # linha[0]=Data, [1]=Nome, [2]=Fralda, [3]=Mimo, [4]=Presença
        if linha[1]: 
            convidado = {
                'data': linha[0],
                'nome': linha[1],
                'fralda': linha[2],
                'mimo': linha[3],
                'presenca': linha[4]
            }
            lista_convidados.append(convidado)
            
    return lista_convidados